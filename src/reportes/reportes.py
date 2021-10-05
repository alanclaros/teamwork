from django.shortcuts import render
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages

# settings de la app
from django.conf import settings
from django.http import HttpResponseRedirect
from django.apps import apps

# propios
from configuraciones.models import Cajas, Puntos, Sucursales, Ciudades
from status.models import Status

# para los usuarios
from utils.permissions import current_periodo, get_user_permission_operation, get_permissions_user, rango_periodos
from utils.dates_functions import get_date_system, get_date_show, get_date_to_db

# clases por modulo
from controllers.reportes.ReportesController import ReportesController
from controllers.ListasController import ListasController

import os
# xls
import openpyxl

# reportes
import io
from django.http import FileResponse, HttpResponse

from reportes.rptArqueoCaja import rptArqueoCaja
from reportes.rptIngresosCaja import rptIngresosCaja
from reportes.rptEgresosCaja import rptEgresosCaja
from reportes.rptMovimientosCaja import rptMovimientosCaja

from reportes.rptCobros import rptCobros
from reportes.rptCobrosMensuales import rptCobrosMensuales
from reportes.rptCobrosManuales import rptCobrosManuales
from reportes.rptLecturas import rptLecturas
from reportes.rptExpensas import rptExpensas
from reportes.rptCobrosPendientes import rptCobrosPendientes


lista_controller = ListasController()


@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def reportes_index(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = reportes_controller.status_activo

    # operaciones
    if 'operation_x' in request.POST.keys():
        # verificamos operacion valida
        operation = request.POST['operation_x']
        if not operation in ['', 'arqueo_caja', 'ingresos_caja', 'egresos_caja', 'movimientos_caja', 'buscar_sucursal', 'buscar_solo_sucursal', 'buscar_sucursal_punto', 'buscar_caja', 'buscar_punto',
                             'cobros', 'cobros_mensuales', 'cobros_manuales', 'lecturas', 'expensas', 'cobros_pendientes']:

            return render(request, 'pages/without_permission.html', {})

        # cobros
        if operation == 'cobros':
            respuesta = cobros(request)
            if not type(respuesta) == bool:
                return respuesta

        # cobros mensuales
        if operation == 'cobros_mensuales':
            respuesta = cobros_mensuales(request)
            if not type(respuesta) == bool:
                return respuesta

        # cobros manuales
        if operation == 'cobros_manuales':
            respuesta = cobros_manuales(request)
            if not type(respuesta) == bool:
                return respuesta

        # lecturas
        if operation == 'lecturas':
            respuesta = lecturas(request)
            if not type(respuesta) == bool:
                return respuesta

        # expensas
        if operation == 'expensas':
            respuesta = expensas(request)
            if not type(respuesta) == bool:
                return respuesta

        # cobros_pendientes
        if operation == 'cobros_pendientes':
            respuesta = cobros_pendientes(request)
            if not type(respuesta) == bool:
                return respuesta

        # cajas
        if operation == 'movimientos_caja':
            respuesta = movimientos_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'egresos_caja':
            respuesta = egresos_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'ingresos_caja':
            respuesta = ingresos_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'arqueo_caja':
            respuesta = arqueo_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'buscar_sucursal':
            ciudad_id = request.POST['ciudad'].strip()
            ciudad = Ciudades.objects.get(pk=ciudad_id)
            lista_sucursales = Sucursales.objects.filter(status_id=status_activo, ciudad_id=ciudad).order_by('sucursal')

            context_s = {
                'lista_sucursales': lista_sucursales,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_sucursal.html', context_s)

        if operation == 'buscar_solo_sucursal':
            ciudad_id = request.POST['ciudad'].strip()
            ciudad = Ciudades.objects.get(pk=ciudad_id)
            lista_sucursales = Sucursales.objects.filter(status_id=status_activo, ciudad_id=ciudad).order_by('sucursal')

            context_s = {
                'lista_sucursales': lista_sucursales,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_solo_sucursal.html', context_s)

        if operation == 'buscar_sucursal_punto':
            ciudad_id = request.POST['ciudad'].strip()
            ciudad = Ciudades.objects.get(pk=ciudad_id)
            lista_sucursales = Sucursales.objects.filter(status_id=status_activo, ciudad_id=ciudad).order_by('sucursal')

            context_s = {
                'lista_sucursales': lista_sucursales,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_sucursal_punto.html', context_s)

        if operation == 'buscar_caja':
            sucursal_id = request.POST['sucursal'].strip()
            sucursal = Sucursales.objects.get(pk=sucursal_id)
            filtro = {}
            filtro['status_id'] = status_activo
            filtro['punto_id__sucursal_id'] = sucursal

            lista_cajas = Cajas.objects.select_related('punto_id').filter(**filtro).order_by('punto_id__punto', 'caja')

            context_s = {
                'lista_cajas': lista_cajas,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_caja.html', context_s)

        if operation == 'buscar_punto':
            sucursal_id = request.POST['sucursal'].strip()
            sucursal = Sucursales.objects.get(pk=sucursal_id)
            filtro = {}
            filtro['status_id'] = status_activo
            filtro['sucursal_id'] = sucursal

            lista_puntos = Puntos.objects.select_related('sucursal_id').filter(**filtro).order_by('punto')

            context_s = {
                'lista_puntos': lista_puntos,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_punto.html', context_s)

    # verificamos mensajes
    if 'nuevo_mensaje' in request.session.keys():
        messages.add_message(request, messages.SUCCESS, request.session['nuevo_mensaje'])
        del request.session['nuevo_mensaje']
        request.session.modified = True

    # datos por defecto
    context = {
        'permisos': permisos,
        'url_main': '',
        # 'js_file': reportes_controller.modulo_session,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': '',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/reportes.html', context)


# arqueo de caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def arqueo_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja
        fecha = get_date_to_db(fecha=request.POST['fecha'], formato_ori='dd-MMM-yyyy', formato='yyyy-mm-dd')

        if permisos.imprimir and reportes_controller.verificar_permiso(request.user, caja_id=int(request.POST['id'])):
            buffer = io.BytesIO()
            rptArqueoCaja(buffer, int(request.POST['id']), fecha)

            buffer.seek(0)
            return FileResponse(buffer, filename='arqueo_caja.pdf')
        else:
            return render(request, 'pages/without_permission.html', {})

    # lista de cajas de la sucursal
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'arqueo_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/arqueo_caja.html', context)


# ingresos a caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def ingresos_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptIngresosCaja(buffer, request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            return FileResponse(buffer, filename='ingresos_caja.pdf')
        else:
            return render(request, 'pages/without_permission.html', {})

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Ciudad', 'Sucursal', 'Punto', 'Caja', 'Operacion', 'Fecha', 'concepto', 'monto', 'codigo', 'estado'))
            datos_reporte = reportes_controller.datos_ingreso_caja(request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['punto'], dato['caja'], dato['operacion'], dato['fecha'], dato['concepto'], dato['monto'], dato['tipo_moneda'], dato['estado_txt']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=ingresos_caja.xlsx"
            wb.save(response)
            return response
        else:
            return render(request, 'pages/without_permission.html', {})

    # lista ciudades
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    # lista de cajas de la sucursal
    #lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'fecha_actual': fecha_actual,
        'lista_ciudades': lista_ciudades,
        'permisos': permisos,
        'url_actual': '',

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'ingresos_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/ingresos_caja.html', context)


# egresos caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def egresos_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptEgresosCaja(buffer, request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='egresos_caja.pdf')
        else:
            return render(request, 'pages/without_permission.html', {})

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Ciudad', 'Sucursal', 'Punto', 'Caja', 'Operacion', 'Fecha', 'concepto', 'monto', 'codigo', 'estado'))
            datos_reporte = reportes_controller.datos_egreso_caja(request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['punto'], dato['caja'], dato['operacion'], dato['fecha'], dato['concepto'], dato['monto'], dato['tipo_moneda'], dato['estado_txt']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=egresos_caja.xlsx"
            wb.save(response)
            return response
        else:
            return render(request, 'pages/without_permission.html', {})

    # lista ciudades
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    # lista de cajas
    #lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'egresos_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/egresos_caja.html', context)


# movimientos caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def movimientos_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptMovimientosCaja(buffer, request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='movimientos_caja.pdf')
        else:
            return render(request, 'pages/without_permission.html', {})

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Ciudad', 'Sucursal', 'Fecha', 'Concepto', 'Monto', 'Codigo', 'Caja1', 'Punto1', 'Caja2', 'Punto2', 'estado'))
            datos_reporte = reportes_controller.datos_movimientos_caja(request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['fecha'], dato['concepto'], dato['monto'], dato['tipo_moneda'], dato['caja1'], dato['punto1'], dato['caja2'], dato['punto2'], dato['estado_txt']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=movimientos_caja.xlsx"
            wb.save(response)
            return response
        else:
            return render(request, 'pages/without_permission.html', {})

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    #lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'movimientos_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/movimientos_caja.html', context)


# cobros
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def cobros(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptCobros(buffer, request.user, bloque, piso, caja, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            return FileResponse(buffer, filename='cobros_realizados.pdf')
        else:
            return render(request, 'pages/without_permission.html', {})

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Caja', 'Monto Bs', 'Periodo', 'Fecha Cobro', 'Bloque', 'Piso', 'Departamento', 'anulacion'))
            datos_reporte = reportes_controller.datos_cobros(request.user, bloque, piso, caja, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['caja'], dato['monto_bs'], dato['periodo'], dato['fecha_cobro'], dato['bloque'], dato['piso'], dato['departamento'], dato['anulacion']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=cobros_realizados.xlsx"
            wb.save(response)
            return response
        else:
            return render(request, 'pages/without_permission.html', {})

    # lista bloques
    lista_bloques = lista_controller.get_lista_bloques(request.user, settings.MOD_REPORTES)
    lista_pisos = lista_controller.get_lista_pisos(request.user, settings.MOD_REPORTES)

    # lista de cajas de la sucursal
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'fecha_actual': fecha_actual,
        'lista_bloques': lista_bloques,
        'lista_pisos': lista_pisos,
        'permisos': permisos,
        'url_actual': '',

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'cobros',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/cobros.html', context)


# cobros mensuales
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def cobros_mensuales(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())
        cobro_mensual = int(request.POST['cobro_mensual'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptCobrosMensuales(buffer, request.user, bloque, piso, caja, fecha_ini, fecha_fin, anulados, cobro_mensual)

            buffer.seek(0)
            return FileResponse(buffer, filename='cobros_mensuales_realizados.pdf')
        else:
            return render(request, 'pages/without_permission.html', {})

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())
        cobro_mensual = int(request.POST['cobro_mensual'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Caja', 'Monto Bs', 'Periodo', 'Fecha Cobro', 'Cobro Mensual', 'Bloque', 'Piso', 'Departamento', 'anulacion'))
            datos_reporte = reportes_controller.datos_cobros_mensuales(request.user, bloque, piso, caja, fecha_ini, fecha_fin, anulados, cobro_mensual)
            for dato in datos_reporte:
                hoja.append((dato['caja'], dato['monto_bs'], dato['periodo'], dato['fecha_cobro'], dato['cobro_mensual'], dato['bloque'], dato['piso'], dato['departamento'], dato['anulacion']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=cobros_mensuales_realizados.xlsx"
            wb.save(response)
            return response
        else:
            return render(request, 'pages/without_permission.html', {})

    # lista bloques
    lista_bloques = lista_controller.get_lista_bloques(request.user, settings.MOD_REPORTES)
    lista_pisos = lista_controller.get_lista_pisos(request.user, settings.MOD_REPORTES)
    lista_mensuales = lista_controller.get_lista_cobros_mensuales(request.user, settings.MOD_REPORTES)

    # lista de cajas de la sucursal
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'fecha_actual': fecha_actual,
        'lista_bloques': lista_bloques,
        'lista_pisos': lista_pisos,
        'lista_mensuales': lista_mensuales,
        'permisos': permisos,
        'url_actual': '',

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'cobros_mensuales',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/cobros_mensuales.html', context)


# cobros manuales
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def cobros_manuales(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())
        cobro_manual = int(request.POST['cobro_manual'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptCobrosManuales(buffer, request.user, bloque, piso, caja, fecha_ini, fecha_fin, anulados, cobro_manual)

            buffer.seek(0)
            return FileResponse(buffer, filename='cobros_manuales_realizados.pdf')
        else:
            return render(request, 'pages/without_permission.html', {})

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())
        cobro_manual = int(request.POST['cobro_manual'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Caja', 'Monto Bs', 'Periodo', 'Fecha Cobro', 'Cobro Manual', 'Bloque', 'Piso', 'Departamento', 'anulacion'))
            datos_reporte = reportes_controller.datos_cobros_manuales(request.user, bloque, piso, caja, fecha_ini, fecha_fin, anulados, cobro_manual)
            for dato in datos_reporte:
                hoja.append((dato['caja'], dato['monto_bs'], dato['periodo'], dato['fecha_cobro'], dato['cobro_manual'], dato['bloque'], dato['piso'], dato['departamento'], dato['anulacion']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=cobros_manuales_realizados.xlsx"
            wb.save(response)
            return response
        else:
            return render(request, 'pages/without_permission.html', {})

    # lista bloques
    lista_bloques = lista_controller.get_lista_bloques(request.user, settings.MOD_REPORTES)
    lista_pisos = lista_controller.get_lista_pisos(request.user, settings.MOD_REPORTES)
    lista_manuales = lista_controller.get_lista_cobros_manuales(request.user, settings.MOD_REPORTES)

    # lista de cajas de la sucursal
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'fecha_actual': fecha_actual,
        'lista_bloques': lista_bloques,
        'lista_pisos': lista_pisos,
        'lista_manuales': lista_manuales,
        'permisos': permisos,
        'url_actual': '',

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'cobros_manuales',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/cobros_manuales.html', context)


# lecturas
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def lecturas(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        periodo_ini = request.POST['periodo_ini'].strip()
        periodo_fin = request.POST['periodo_fin'].strip()
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptLecturas(buffer, request.user, bloque, piso, caja, periodo_ini, periodo_fin, anulados)

            buffer.seek(0)
            return FileResponse(buffer, filename='lecturas_realizados.pdf')
        else:
            return render(request, 'pages/without_permission.html', {})

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        periodo_ini = request.POST['periodo_ini'].strip()
        periodo_fin = request.POST['periodo_fin'].strip()
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Caja', 'Monto Bs', 'Periodo', 'Fecha Cobro', 'Lectura', 'Bloque', 'Piso', 'Departamento', 'anulacion'))
            datos_reporte = reportes_controller.datos_lecturas(request.user, bloque, piso, caja, periodo_ini, periodo_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['caja'], dato['consumo'], dato['periodo'], dato['fecha_cobro'], dato['lectura'], dato['bloque'], dato['piso'], dato['departamento'], dato['anulacion']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=lecturas_realizados.xlsx"
            wb.save(response)
            return response
        else:
            return render(request, 'pages/without_permission.html', {})

    # lista bloques
    lista_bloques = lista_controller.get_lista_bloques(request.user, settings.MOD_REPORTES)
    lista_pisos = lista_controller.get_lista_pisos(request.user, settings.MOD_REPORTES)

    periodo_actual = current_periodo()
    lista_periodos = rango_periodos(periodo_actual)

    # lista de cajas de la sucursal
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'fecha_actual': fecha_actual,
        'lista_bloques': lista_bloques,
        'lista_pisos': lista_pisos,
        'lista_periodos': lista_periodos,
        'permisos': permisos,
        'url_actual': '',

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'lecturas',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/lecturas.html', context)


# expensas
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def expensas(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        periodo_ini = request.POST['periodo_ini'].strip()
        periodo_fin = request.POST['periodo_fin'].strip()
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptExpensas(buffer, request.user, bloque, piso, caja, periodo_ini, periodo_fin, anulados)

            buffer.seek(0)
            return FileResponse(buffer, filename='expensas_realizados.pdf')
        else:
            return render(request, 'pages/without_permission.html', {})

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        periodo_ini = request.POST['periodo_ini'].strip()
        periodo_fin = request.POST['periodo_fin'].strip()
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Caja', 'Monto Bs', 'Periodo', 'Fecha Cobro', 'Lectura', 'Bloque', 'Piso', 'Departamento', 'anulacion'))
            datos_reporte = reportes_controller.datos_expensas(request.user, bloque, piso, caja, periodo_ini, periodo_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['caja'], dato['total_expensas'], dato['periodo'], dato['fecha_cobro'], dato['lectura'], dato['bloque'], dato['piso'], dato['departamento'], dato['anulacion']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=expensas_realizados.xlsx"
            wb.save(response)
            return response
        else:
            return render(request, 'pages/without_permission.html', {})

    # lista bloques
    lista_bloques = lista_controller.get_lista_bloques(request.user, settings.MOD_REPORTES)
    lista_pisos = lista_controller.get_lista_pisos(request.user, settings.MOD_REPORTES)

    periodo_actual = current_periodo()
    lista_periodos = rango_periodos(periodo_actual)

    # lista de cajas de la sucursal
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'fecha_actual': fecha_actual,
        'lista_bloques': lista_bloques,
        'lista_pisos': lista_pisos,
        'lista_periodos': lista_periodos,
        'permisos': permisos,
        'url_actual': '',

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'expensas',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/expensas.html', context)


# cobros pendientes
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def cobros_pendientes(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        periodo_ini = request.POST['periodo_ini'].strip()
        periodo_fin = request.POST['periodo_fin'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptCobrosPendientes(buffer, request.user, bloque, piso, periodo_ini, periodo_fin)

            buffer.seek(0)
            return FileResponse(buffer, filename='cobros_pendientes.pdf')
        else:
            return render(request, 'pages/without_permission.html', {})

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        periodo_ini = request.POST['periodo_ini'].strip()
        periodo_fin = request.POST['periodo_fin'].strip()
        bloque = int(request.POST['bloque'].strip())
        piso = int(request.POST['piso'].strip())

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Periodo', 'Bloque', 'Piso', 'Departamento', 'Pendiente'))
            datos_reporte = reportes_controller.datos_cobros_pendientes(request.user, bloque, piso, periodo_ini, periodo_fin)
            for dato in datos_reporte:
                hoja.append((dato['periodo'], dato['bloque'], dato['piso'], dato['departamento'], dato['monto_bs']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=cobros_pendientes.xlsx"
            wb.save(response)
            return response
        else:
            return render(request, 'pages/without_permission.html', {})

    # lista bloques
    lista_bloques = lista_controller.get_lista_bloques(request.user, settings.MOD_REPORTES)
    lista_pisos = lista_controller.get_lista_pisos(request.user, settings.MOD_REPORTES)

    periodo_actual = current_periodo()
    lista_periodos = rango_periodos(periodo_actual)

    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'fecha_actual': fecha_actual,
        'lista_bloques': lista_bloques,
        'lista_pisos': lista_pisos,
        'lista_periodos': lista_periodos,
        'permisos': permisos,
        'url_actual': '',

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'cobros_pendientes',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/cobros_pendientes.html', context)
